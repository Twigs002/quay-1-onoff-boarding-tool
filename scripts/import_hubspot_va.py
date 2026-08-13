#!/usr/bin/env python3
"""
import_hubspot_va.py - one-time (repeatable) bulk import of a HubSpot CRM contact
export into the VA Searches tracker table `va_search_records` (see
supabase/migrations/0001_va_search_records.sql and web/app.vasearches.js).

Each contact becomes a search row: `name` from First/Last, `existing_number` from
the best phone on record, outcome 'pending' (the VA still has to find/verify a
number). Rich HubSpot fields (ID number, division, suburb, address, lead status,
contact owner) are kept. Dedupe is by HubSpot Record ID stored as
`source_id = 'hubspot:<Record ID>'`, so re-running a fresh export ADDS only new
contacts and never overwrites a VA's already-recorded result.

Entity type (auto by default): a row with no First+Last name (e.g. a body-corporate
line) is tagged 'company'; everything else 'person'. Force with --entity.

Auth: needs the Supabase SERVICE-ROLE key (bypasses RLS) - NEVER the anon key.
Provide via env, e.g.:
    export SUPABASE_URL='https://dqszbqiimbfvmmnpgpsb.supabase.co'
    export SUPABASE_SERVICE_KEY='...service role key...'

Safety: defaults to --dry-run (parses + prints a summary + sample rows, writes
nothing). Pass --live to actually upsert. Requires: pip install openpyxl

Usage:
    python3 scripts/import_hubspot_va.py "/path/export.xlsx"                 # dry run
    python3 scripts/import_hubspot_va.py "/path/export.xlsx" --sheet "HubSpot 2026-08-12"
    python3 scripts/import_hubspot_va.py "/path/export.xlsx" --entity person --live
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

BATCH = 500
TABLE = "va_search_records"
# Header labels as they appear in the HubSpot "Data View" export. Matched by name
# (case-insensitive), so column order can change without breaking the import.
COL = {
    "record_id": "Record ID",
    "first": "First Name",
    "last": "Last Name",
    "unit": "Unit / Street number",
    "division": "Division",
    "lead_status": "Lead Status",
    "suburb": "Suburb",
    "street": "Block / Street",
    "owner": "Contact owner",
    "id_number": "ID Number",
    # phones, best-first: a dedicated Phone Number wins over the mobiles.
    "phones": ["Phone Number", "Mobile Phone Number", "Mobile phone 2",
               "Mobile phone 3", "Mobile phone 4"],
}
# Placeholder tokens HubSpot uses for "nothing here".
BLANKS = {"", "na", "n/a", "none", "null", "-", "0"}


def clean(v) -> str:
    return "" if v is None else str(v).strip()


def is_blank(v: str) -> bool:
    return clean(v).lower() in BLANKS


def looks_like_number(v: str) -> bool:
    return len(re.sub(r"\D", "", clean(v))) >= 7


def header_index(header: list[str]) -> dict[str, int]:
    idx = {}
    for i, h in enumerate(header):
        idx[clean(h).lower()] = i
    return idx


def col_i(hidx: dict[str, int], label: str) -> int | None:
    return hidx.get(label.lower())


def build_record(row: list, hidx: dict, sheet: str, entity_mode: str) -> dict | None:
    def get(label: str) -> str:
        i = col_i(hidx, label)
        return clean(row[i]) if (i is not None and i < len(row)) else ""

    first, last = get(COL["first"]), get(COL["last"])
    name = " ".join(p for p in (first, last) if p).strip()

    # Best phone on record, plus any extras for the notes line.
    numbers = []
    for lbl in COL["phones"]:
        val = get(lbl)
        if not is_blank(val) and looks_like_number(val) and val not in numbers:
            numbers.append(val)
    existing_number = numbers[0] if numbers else None
    extra_numbers = numbers[1:]

    unit, street, suburb = get(COL["unit"]), get(COL["street"]), get(COL["suburb"])
    address = " ".join(p for p in (unit, street) if p).strip()
    if suburb and address:
        address = f"{address}, {suburb}"
    elif suburb:
        address = suburb

    division, lead_status = get(COL["division"]), get(COL["lead_status"])
    owner, id_number = get(COL["owner"]), get(COL["id_number"])
    record_id = get(COL["record_id"])

    # Entity: auto -> blank person-name means it's a company/body-corporate line.
    if entity_mode == "auto":
        entity_type = "person" if name else "company"
    else:
        entity_type = entity_mode

    # A blank-name row still needs a label: fall back to address, then division,
    # then the record id, so it is findable in the log.
    if not name:
        name = address or division or (f"Record {record_id}" if record_id else "(unnamed)")

    notes = None
    if extra_numbers:
        notes = "Other numbers on record: " + ", ".join(extra_numbers)

    # A contact that ALREADY has a number on record is not a blank "pending" job -
    # it counts as found (same number we already had), so the dashboard reflects
    # real progress. Only contacts with no number import as pending to be searched.
    if existing_number:
        outcome, found_number, searched_by = "found_unchanged", existing_number, "HubSpot import"
    else:
        outcome, found_number, searched_by = "pending", None, None

    rec = {
        "entity_type": entity_type,
        "name": name,
        "sheet": sheet,
        "existing_number": existing_number,
        "found_number": found_number,
        "outcome": outcome,
        "searched_by": searched_by,
        "id_number": id_number or None,
        "division": division or None,
        "suburb": suburb or None,
        "address": address or None,
        "lead_status": lead_status or None,
        "contact_owner": owner or None,
        "source_id": f"hubspot:{record_id}" if record_id else None,
        "notes": notes,
        "created_by": "HubSpot import",
    }
    return rec


def load_rows(path: Path, sheet: str, entity_mode: str, limit: int | None):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(it)]
    hidx = header_index(header)
    missing = [COL["record_id"], COL["first"], COL["last"]]
    for req in missing:
        if col_i(hidx, req) is None:
            sys.exit(f"Expected column '{req}' not found. Headers seen: {header}")

    records, seen_src, skipped_dupe = [], set(), 0
    for row in it:
        rec = build_record(list(row), hidx, sheet, entity_mode)
        if rec is None:
            continue
        src = rec.get("source_id")
        if src:
            if src in seen_src:      # dedupe WITHIN the file (same batch can't have 2)
                skipped_dupe += 1
                continue
            seen_src.add(src)
        records.append(rec)
        if limit and len(records) >= limit:
            break
    return records, skipped_dupe


def upsert(records: list[dict]) -> None:
    url = os.environ.get("SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("Set SUPABASE_URL and SUPABASE_SERVICE_KEY (service-role key) in the environment.")
    endpoint = f"{url}/rest/v1/{TABLE}?on_conflict=source_id"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        # Insert; on an existing source_id, do nothing (preserve VA progress).
        "Prefer": "resolution=ignore-duplicates,return=minimal",
    }
    total = len(records)
    for start in range(0, total, BATCH):
        chunk = records[start:start + BATCH]
        body = json.dumps(chunk).encode("utf-8")
        req = urllib.request.Request(endpoint, data=body, headers=headers, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                resp.read()
        except urllib.error.HTTPError as e:
            detail = e.read().decode("utf-8", "replace")[:500]
            sys.exit(f"\nHTTP {e.code} on rows {start}-{start + len(chunk)}: {detail}")
        except urllib.error.URLError as e:
            sys.exit(f"\nNetwork error on rows {start}-{start + len(chunk)}: {e}")
        print(f"  upserted {min(start + BATCH, total)}/{total}", flush=True)


def main() -> None:
    ap = argparse.ArgumentParser(description="Bulk import a HubSpot CRM export into va_search_records.")
    ap.add_argument("xlsx", help="Path to the HubSpot .xlsx export")
    ap.add_argument("--sheet", default=None, help="Batch/sheet label (default derived from filename date)")
    ap.add_argument("--entity", choices=["auto", "person", "company"], default="auto",
                    help="Entity type for imported rows (default: auto-detect by blank name)")
    ap.add_argument("--limit", type=int, default=None, help="Only import the first N rows (testing)")
    ap.add_argument("--live", action="store_true", help="Actually write. Omit for a dry run.")
    ap.add_argument("--dry-run", action="store_true", help="Parse + summarise only (default).")
    args = ap.parse_args()

    path = Path(args.xlsx).expanduser()
    if not path.exists():
        sys.exit(f"File not found: {path}")

    # Default sheet label from a YYYY-MM-DD in the filename, else the stem.
    sheet = args.sheet
    if not sheet:
        m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
        sheet = f"HubSpot {m.group(1)}" if m else f"HubSpot {path.stem}"[:60]

    records, skipped_dupe = load_rows(path, sheet, args.entity, args.limit)
    n_company = sum(1 for r in records if r["entity_type"] == "company")
    n_person = len(records) - n_company
    n_with_num = sum(1 for r in records if r["existing_number"])

    print(f"Parsed {len(records)} rows from {path.name}")
    print(f"  sheet label     : {sheet!r}")
    print(f"  person / company: {n_person} / {n_company}")
    print(f"  already found   : {n_with_num}  (imported as found - number already on record)")
    print(f"  to search       : {len(records) - n_with_num}  (imported as pending - no number yet)")
    if skipped_dupe:
        print(f"  in-file dupes skipped (same Record ID): {skipped_dupe}")
    print("  sample rows:")
    for r in records[:3]:
        print("   ", {k: r[k] for k in ("entity_type", "name", "existing_number", "id_number", "division", "source_id")})

    live = args.live and not args.dry_run
    if not live:
        print("\nDRY RUN - nothing written. Re-run with --live to upsert.")
        return
    print(f"\nLIVE: upserting {len(records)} rows (ignore-duplicates on source_id)...")
    upsert(records)
    print("Done.")


if __name__ == "__main__":
    main()
