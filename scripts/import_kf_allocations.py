#!/usr/bin/env python3
"""
import_kf_allocations.py - load "MASTER KF Team Allocations.xlsx" into the
`kf_allocations` reference table (quay-clock Supabase) for the hub's Allocations tab.

Reads the FT + ST master sheets (title_type FT/ST) and the New Request sheet
(status=requested). Upserts on a stable source_key so re-running a fresh master
updates in place instead of duplicating.

Auth: SERVICE-ROLE key (bypasses RLS) - env SUPABASE_SERVICE_KEY, or Keychain
'supabase-va'. URL defaults to quay-clock. Requires: pip install openpyxl.

Usage:
    python3 scripts/import_kf_allocations.py "~/Downloads/MASTER KF Team Allocations.xlsx"          # dry run
    python3 scripts/import_kf_allocations.py "~/Downloads/MASTER KF Team Allocations.xlsx" --live   # write
"""
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import urllib.error
import urllib.request
from pathlib import Path

DEFAULT_URL = "https://dqszbqiimbfvmmnpgpsb.supabase.co"
TABLE = "kf_allocations"
BATCH = 500
STATUS_MAP = {"active": "active", "on ice": "on_ice", "onice": "on_ice",
              "transferred": "transferred", "inactive": "inactive"}


def clean(v) -> str:
    return "" if v is None else str(v).strip()


def norm_owner(v: str) -> str:
    v = clean(v)
    return v[:-2] if v.endswith(".0") else v


def creds() -> tuple[str, str]:
    url = (os.environ.get("SUPABASE_URL") or DEFAULT_URL).rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not key:
        try:
            key = subprocess.check_output(
                ["security", "find-generic-password", "-s", "supabase-va", "-a", "dqszbqiimbfvmmnpgpsb", "-w"],
                text=True, stderr=subprocess.DEVNULL).strip()
        except Exception:
            key = ""
    if not key:
        sys.exit("No Supabase service key (env SUPABASE_SERVICE_KEY or Keychain 'supabase-va').")
    return url, key


def rows_from(ws, title_type: str | None, source: str, is_request: bool) -> list[dict]:
    it = ws.iter_rows(values_only=True)
    next(it, None)  # header
    out, seen = [], {}
    for r in it:
        if not any(x is not None and clean(x) for x in r):
            continue
        ext_name, ext, aka = clean(r[0]), clean(r[1]), clean(r[2])
        suburb = aka or ext_name
        if not suburb:
            continue
        if is_request:
            rec = {"title_type": None, "suburb": suburb, "extension_name": ext_name, "extension": ext,
                   "team": None, "status": "requested", "request_period": clean(r[3]) or None, "source": source}
        else:
            team = clean(r[3])
            rec = {"title_type": title_type, "suburb": suburb, "extension_name": ext_name, "extension": ext,
                   "cma_suburb": clean(r[5]) or None, "team": team or None, "owner_id": norm_owner(r[4]) or None,
                   "status": STATUS_MAP.get(clean(r[6]).lower(), "active"),
                   "original_accountability": clean(r[7]) or None, "last_transferred": clean(r[8]) or None,
                   "rt_period": clean(r[9]) or None, "request_period": clean(r[10]) or None, "source": source}
        base = f"{title_type or 'REQ'}|{(rec.get('cma_suburb') or suburb).lower()}|{(rec.get('team') or '').lower()}"
        n = seen.get(base, 0); seen[base] = n + 1
        rec["source_key"] = base if n == 0 else f"{base}#{n}"
        rec["created_by"] = "KF master import"
        out.append(rec)
    return out


def upsert(url: str, key: str, rows: list[dict]) -> None:
    hdr = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json",
           "Prefer": "resolution=merge-duplicates,return=minimal"}
    ep = f"{url}/rest/v1/{TABLE}?on_conflict=source_key"
    for i in range(0, len(rows), BATCH):
        chunk = rows[i:i + BATCH]
        req = urllib.request.Request(ep, data=json.dumps(chunk).encode(), headers=hdr, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=60) as r:
                r.read()
        except urllib.error.HTTPError as e:
            sys.exit(f"HTTP {e.code} on rows {i}-{i+len(chunk)}: {e.read().decode('utf-8','replace')[:400]}")
        print(f"  upserted {min(i+BATCH, len(rows))}/{len(rows)}", flush=True)


def main() -> None:
    ap = argparse.ArgumentParser(description="Import MASTER KF Team Allocations into kf_allocations.")
    ap.add_argument("xlsx")
    ap.add_argument("--live", action="store_true", help="Write (default: dry run).")
    args = ap.parse_args()
    path = Path(args.xlsx).expanduser()
    if not path.exists():
        sys.exit(f"File not found: {path}")
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    rows: list[dict] = []
    plan = [("MASTER FT ALLOCATIONS", "FT", "FT master", False),
            ("MASTER ST ALLOCATIONS", "ST", "ST master", False),
            ("New Request", None, "new request", True)]
    for sheet, tt, src, isreq in plan:
        if sheet in wb.sheetnames:
            got = rows_from(wb[sheet], tt, src, isreq)
            rows += got
            print(f"  {sheet}: {len(got)} rows")

    by_status: dict[str, int] = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"\nParsed {len(rows)} allocations · by status: {by_status}")
    print("  sample:", {k: rows[0].get(k) for k in ("title_type", "suburb", "team", "status", "cma_suburb")} if rows else {})

    if not args.live:
        print("\nDRY RUN - nothing written. Re-run with --live to upsert.")
        return
    url, key = creds()
    print(f"\nLIVE: upserting {len(rows)} rows...")
    upsert(url, key, rows)
    print("Done.")


if __name__ == "__main__":
    main()
